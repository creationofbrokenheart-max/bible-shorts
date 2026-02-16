import json
from openpyxl import load_workbook
from pathlib import Path

XLSX_PATH = Path("verses.xlsx")
SHEET_NAME = "Sheet2"   # adjust if different
OUTPUT_JSON = Path("current_verse.json")

def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    # Find header columns
    header_row = 1
    headers = {cell.value: cell.column for cell in ws[header_row]}
    ref_col = headers.get("Verus")
    status_col = headers.get("Status")
    if ref_col is None or status_col is None:
        raise ValueError("Columns 'Verus' or 'Status' not found in sheet.")

    selected = None
    selected_row_idx = None

    # Iterate rows after header
    for row in range(header_row + 1, ws.max_row + 1):
        ref_val = ws.cell(row=row, column=ref_col).value
        status_val = ws.cell(row=row, column=status_col).value

        if ref_val and (status_val is None or str(status_val).strip() == ""):
            selected = {
                "row_index": row,
                "reference": str(ref_val).strip(),
                "status": None,
            }
            selected_row_idx = row
            break

    if not selected:
        print("No verse with blank Status found.")
        return

    # Save for downstream scripts
    OUTPUT_JSON.write_text(json.dumps(selected, ensure_ascii=False, indent=2))
    print(f"Selected verse: {selected['reference']} (row {selected_row_idx})")

if __name__ == "__main__":
    main()
